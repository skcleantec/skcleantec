"""청소비서 API — 로그인 · 엑셀 일괄 접수 미리보기/실행."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import requests

DEFAULT_API_BASE = 'https://www.cbiseo.com'


class CbiseoApiError(Exception):
    def __init__(self, message: str, *, status_code: int | None = None) -> None:
        super().__init__(message)
        self.status_code = status_code


DEFAULT_BATCH_SIZE = 40


class CbiseoClient:
    def __init__(self, base_url: str = DEFAULT_API_BASE, timeout: int = 120) -> None:
        self.base_url = base_url.rstrip('/')
        self.timeout = timeout
        self.token: str | None = None
        self.user: dict[str, Any] | None = None

    def _headers(self, *, json_body: bool = False) -> dict[str, str]:
        headers: dict[str, str] = {}
        if json_body:
            headers['Content-Type'] = 'application/json'
        if self.token:
            headers['Authorization'] = f'Bearer {self.token}'
        return headers

    def _parse_error(self, response: requests.Response, fallback: str) -> str:
        try:
            data = response.json()
            if isinstance(data, dict) and data.get('error'):
                return str(data['error'])
        except (ValueError, json.JSONDecodeError):
            pass
        return fallback

    def login(self, tenant_slug: str, email: str, password: str) -> dict[str, Any]:
        url = f'{self.base_url}/api/auth/login'
        payload = {
            'tenantSlug': tenant_slug.strip(),
            'email': email.strip().lower(),
            'password': password,
        }
        response = requests.post(url, json=payload, timeout=self.timeout)
        if not response.ok:
            raise CbiseoApiError(
                self._parse_error(response, '로그인에 실패했습니다.'),
                status_code=response.status_code,
            )
        data = response.json()
        token = data.get('token')
        if not token:
            raise CbiseoApiError('로그인 응답에 토큰이 없습니다.')
        self.token = str(token)
        self.user = data.get('user') if isinstance(data.get('user'), dict) else None
        return data

    def get_field_catalog(self) -> list[dict[str, Any]]:
        self._require_token()
        url = f'{self.base_url}/api/inquiry-excel-import/field-catalog'
        response = requests.get(url, headers=self._headers(), timeout=self.timeout)
        if not response.ok:
            raise CbiseoApiError(
                self._parse_error(response, '필드 목록을 불러올 수 없습니다.'),
                status_code=response.status_code,
            )
        data = response.json()
        return data if isinstance(data, list) else []

    def list_profiles(self) -> list[dict[str, Any]]:
        self._require_token()
        url = f'{self.base_url}/api/inquiry-excel-import/profiles'
        response = requests.get(url, headers=self._headers(), timeout=self.timeout)
        if not response.ok:
            raise CbiseoApiError(
                self._parse_error(response, '매칭 서식 목록을 불러올 수 없습니다.'),
                status_code=response.status_code,
            )
        data = response.json()
        items = data.get('items') if isinstance(data, dict) else []
        return items if isinstance(items, list) else []

    def update_profile(self, profile_id: str, mapping_spec: dict[str, Any]) -> dict[str, Any]:
        self._require_token()
        url = f'{self.base_url}/api/inquiry-excel-import/profiles/{profile_id}'
        response = requests.patch(
            url,
            headers=self._headers(json_body=True),
            json={'mappingSpec': mapping_spec},
            timeout=self.timeout,
        )
        if not response.ok:
            raise CbiseoApiError(
                self._parse_error(response, '매칭 서식 수정에 실패했습니다.'),
                status_code=response.status_code,
            )
        return response.json()

    def create_profile(self, name: str, mapping_spec: dict[str, Any]) -> dict[str, Any]:
        self._require_token()
        url = f'{self.base_url}/api/inquiry-excel-import/profiles'
        response = requests.post(
            url,
            headers=self._headers(json_body=True),
            json={'name': name, 'mappingSpec': mapping_spec},
            timeout=self.timeout,
        )
        if not response.ok:
            raise CbiseoApiError(
                self._parse_error(response, '매칭 서식 생성에 실패했습니다.'),
                status_code=response.status_code,
            )
        return response.json()

    def find_or_create_timetree_profile(self, mapping_spec: dict[str, Any]) -> dict[str, Any]:
        target_name = '타임트리 자동'
        for profile in self.list_profiles():
            if str(profile.get('name', '')).strip() == target_name:
                profile_id = str(profile.get('id', '')).strip()
                if profile_id:
                    return self.update_profile(profile_id, mapping_spec)
                return profile
        return self.create_profile(target_name, mapping_spec)

    def preview_import(self, profile_id: str, xlsx_path: Path) -> dict[str, Any]:
        self._require_token()
        url = f'{self.base_url}/api/inquiry-excel-import/import/preview'
        preview_timeout = max(self.timeout, 1200)
        with xlsx_path.open('rb') as handle:
            response = requests.post(
                url,
                headers=self._headers(),
                files={'file': (xlsx_path.name, handle, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                data={'profileId': profile_id},
                timeout=preview_timeout,
            )
        if not response.ok:
            raise CbiseoApiError(
                self._parse_error(response, '미리보기에 실패했습니다.'),
                status_code=response.status_code,
            )
        return response.json()

    def list_import_runs(self, *, limit: int = 20, offset: int = 0) -> dict[str, Any]:
        self._require_token()
        url = f'{self.base_url}/api/inquiry-excel-import/runs'
        response = requests.get(
            url,
            headers=self._headers(),
            params={'limit': limit, 'offset': offset},
            timeout=self.timeout,
        )
        if not response.ok:
            raise CbiseoApiError(
                self._parse_error(response, '실행 이력을 불러올 수 없습니다.'),
                status_code=response.status_code,
            )
        data = response.json()
        return data if isinstance(data, dict) else {'items': [], 'total': 0}

    def get_import_run(self, run_id: str) -> dict[str, Any]:
        self._require_token()
        url = f'{self.base_url}/api/inquiry-excel-import/runs/{run_id}'
        response = requests.get(url, headers=self._headers(), timeout=self.timeout)
        if not response.ok:
            raise CbiseoApiError(
                self._parse_error(response, '실행 이력을 불러올 수 없습니다.'),
                status_code=response.status_code,
            )
        data = response.json()
        return data if isinstance(data, dict) else {}

    def delete_import_run_inquiries(self, run_id: str, password: str) -> dict[str, Any]:
        self._require_token()
        url = f'{self.base_url}/api/inquiry-excel-import/runs/{run_id}/inquiries/delete'
        response = requests.post(
            url,
            headers=self._headers(json_body=True),
            json={'password': password},
            timeout=max(self.timeout, 1200),
        )
        if not response.ok:
            raise CbiseoApiError(
                self._parse_error(response, '일괄 삭제에 실패했습니다.'),
                status_code=response.status_code,
            )
        data = response.json()
        return data if isinstance(data, dict) else {}

    def execute_import_batch(
        self,
        profile_id: str,
        xlsx_path: Path,
        *,
        run_id: str | None = None,
        start_offset: int = 0,
        batch_size: int = DEFAULT_BATCH_SIZE,
    ) -> dict[str, Any]:
        self._require_token()
        url = f'{self.base_url}/api/inquiry-excel-import/import/execute/batch'
        batch_timeout = max(self.timeout, 600)
        data: dict[str, str | int] = {
            'profileId': profile_id,
            'startOffset': max(0, int(start_offset)),
            'batchSize': max(1, int(batch_size)),
        }
        if run_id:
            data['runId'] = run_id.strip()
        with xlsx_path.open('rb') as handle:
            response = requests.post(
                url,
                headers=self._headers(),
                files={'file': (xlsx_path.name, handle, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                data=data,
                timeout=batch_timeout,
            )
        if not response.ok:
            raise CbiseoApiError(
                self._parse_error(response, '일괄 등록에 실패했습니다.'),
                status_code=response.status_code,
            )
        return response.json()

    def execute_import_batched(
        self,
        profile_id: str,
        xlsx_path: Path,
        *,
        run_id: str | None = None,
        start_offset: int = 0,
        batch_size: int = DEFAULT_BATCH_SIZE,
        on_progress: Any | None = None,
    ) -> dict[str, Any]:
        offset = max(0, int(start_offset))
        current_run_id = (run_id or '').strip() or None
        latest: dict[str, Any] = {}
        while True:
            latest = self.execute_import_batch(
                profile_id,
                xlsx_path,
                run_id=current_run_id,
                start_offset=offset,
                batch_size=batch_size,
            )
            current_run_id = str(latest.get('runId', '')).strip() or current_run_id
            if on_progress:
                on_progress(latest)
            if latest.get('done'):
                break
            offset = int(latest.get('nextOffset', offset))
            if offset < 0:
                break
        return latest

    def execute_import(self, profile_id: str, xlsx_path: Path) -> dict[str, Any]:
        """레거시 — 배치 방식으로 전체 등록."""
        return self.execute_import_batched(profile_id, xlsx_path)

    def _require_token(self) -> None:
        if not self.token:
            raise CbiseoApiError('로그인이 필요합니다.')


def format_preview_summary(result: dict[str, Any]) -> str:
    lines = [
        f"파일: {result.get('fileName', '')}",
        f"총 {result.get('totalRows', 0)}행",
        f"등록 예정: {result.get('createdCount', 0)}건",
        f"건너뜀: {result.get('skippedCount', 0)}건",
        f"오류: {result.get('errorCount', 0)}건",
        '',
    ]
    preview = result.get('preview') or []
    if isinstance(preview, list):
        for item in preview[:20]:
            if not isinstance(item, dict):
                continue
            action = item.get('action', '')
            row_index = item.get('rowIndex', '')
            message = item.get('message', '')
            lines.append(f"  · {row_index}행 [{action}] {message}")
        if len(preview) > 20:
            lines.append(f'  … 외 {len(preview) - 20}건')
    return '\n'.join(lines)


def format_execute_summary(result: dict[str, Any]) -> str:
    status = str(result.get('status', '') or '')
    last_row = result.get('lastProcessedRowIndex')
    pending = result.get('pendingRowCount')
    processed = result.get('processedRowCount')
    lines = [
        f"실행 ID: {result.get('runId', '')}",
    ]
    if status:
        status_label = {
            'RUNNING': '진행 중(중단됨)',
            'FAILED': '실패(부분 저장)',
            'COMPLETED': '완료',
        }.get(status, status)
        lines.append(f'상태: {status_label}')
    lines.extend([
        f"총 {result.get('totalRows', 0)}행",
        f"등록: {result.get('createdCount', 0)}건",
        f"건너뜀: {result.get('skippedCount', 0)}건",
        f"오류: {result.get('errorCount', 0)}건",
    ])
    if processed is not None:
        lines.append(f'처리 기록: {processed}행')
    if last_row:
        lines.append(f'마지막 처리 행: {last_row}')
    if pending not in (None, 0) and status != 'COMPLETED':
        lines.append(f'남은 행: {pending}건 — 「이어서 등록」을 사용하세요')
    lines.append('')
    rows = result.get('rows') or []
    if isinstance(rows, list):
        for item in rows[:20]:
            if not isinstance(item, dict):
                continue
            kind = item.get('kind', '')
            row_index = item.get('rowIndex', '')
            message = item.get('message', '')
            inquiry_number = item.get('inquiryNumber', '')
            suffix = f' ({inquiry_number})' if inquiry_number else ''
            lines.append(f"  · {row_index}행 [{kind}]{suffix} {message}")
        if len(rows) > 20:
            lines.append(f'  … 외 {len(rows) - 20}건')
    return '\n'.join(lines)


def format_delete_summary(result: dict[str, Any]) -> str:
    deleted = int(result.get('deletedCount', 0) or 0)
    not_found = int(result.get('notFoundCount', 0) or 0)
    already = int(result.get('alreadyDeletedCount', 0) or 0)
    attempted = int(result.get('attemptedCount', 0) or 0)
    unresolved = int(result.get('unresolvedRows', 0) or 0)
    missing_id = int(result.get('missingInquiryIdRows', 0) or 0)
    lines = ['일괄 삭제 결과', f'삭제: {deleted}건']
    if attempted:
        lines.append(f'삭제 대상(이력): {attempted}건')
    if not_found:
        lines.append(f'DB에서 찾을 수 없음: {not_found}건')
    if unresolved:
        lines.append(f'접수 ID·번호로 연결 실패: {unresolved}건')
    if missing_id:
        lines.append(f'이력에 접수 ID 없음: {missing_id}건')
    if already:
        lines.append(f'이미 삭제 처리됨: {already}건')
    if deleted == 0 and attempted > 0:
        lines.append('')
        lines.append('실제 접수가 삭제되지 않았습니다. 다른 실행 이력을 선택했거나, 이미 개별 삭제됐을 수 있습니다.')
    return '\n'.join(lines)


def count_xlsx_data_rows(xlsx_path: Path) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        max_row = int(ws.max_row or 0)
        return max(0, max_row - 1)
    finally:
        wb.close()
