"""청소비서 업로드 패널 (통합 앱 ④단계)."""
from __future__ import annotations

import queue
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk

from cbiseo_export import export_cbiseo_workbook, format_export_summary, sanitize_operating_company
from cbiseo_upload import (
    CbiseoClient,
    count_xlsx_data_rows,
    format_delete_summary,
    format_execute_summary,
    format_preview_summary,
)
from local_settings import save_settings
from upload_progress import PasswordConfirmDialog, TaskProgressDialog
from workflow_context import APP_DIR, WorkflowContext
from workflow_paths import brand_dir, resolve_workspace_root

PANEL_TITLE = '청소비서 업로드'


class UploadPanel(ttk.Frame):
    def __init__(self, parent: tk.Misc, ctx: WorkflowContext) -> None:
        super().__init__(parent)
        self.ctx = ctx
        self.root = ctx.root
        self._queue: queue.Queue[tuple[str, object]] = queue.Queue()
        self._worker: threading.Thread | None = None
        self._api_client: CbiseoClient | None = None
        self._profiles: list[dict] = []
        self._import_runs: list[dict] = []
        self._last_preview: dict[str, int] | None = None
        self._progress_dialog: TaskProgressDialog | None = None
        self._progress_tick_job: str | None = None
        self._progress_mode: str = ''

        s = ctx.settings
        self.api_base_var = tk.StringVar(value=s.api_base_url)
        self.tenant_slug_var = tk.StringVar(value=s.tenant_slug)
        self.api_email_var = tk.StringVar(value=s.api_email)
        self.api_password_var = tk.StringVar(value=s.api_password)
        self.remember_api_var = tk.BooleanVar(value=s.remember_api)
        self.profile_var = tk.StringVar(value='')
        self.operating_company_var = tk.StringVar(value=s.operating_company)
        self.import_ready_only_var = tk.BooleanVar(value=s.upload_import_ready_only)
        self.confirmed_only_var = tk.BooleanVar(value=s.upload_confirmed_only)
        self.import_run_var = tk.StringVar(value='')

        self._build_ui()
        self._normalize_operating_company_field()
        self.root.after(100, self._process_queue)

    def _normalize_operating_company_field(self) -> None:
        oc = self.operating_company_var.get().strip()
        slug = self.tenant_slug_var.get().strip()
        if oc and slug and oc.lower() == slug.lower():
            self.operating_company_var.set('')

    def _build_ui(self) -> None:
        ttk.Label(
            self,
            text='① 일정 추출에서 입력한 청소비서 계정을 사용합니다. 필요하면 아래에서 수정할 수 있습니다.',
            wraplength=900,
            foreground='#475569',
        ).pack(anchor='w', padx=12, pady=(12, 6))

        ttk.Label(
            self,
            text='확정한 행을 청소비서에 일괄 등록합니다. 금액은 원화 정수(예: 200000)로 전송됩니다.',
            wraplength=900,
        ).pack(anchor='w', padx=12, pady=(0, 6))

        api_frame = ttk.LabelFrame(self, text='청소비서 로그인')
        api_frame.pack(fill='x', padx=12, pady=6)

        def api_row(label: str, var: tk.StringVar, *, secret: bool = False, width: int = 48) -> None:
            row = ttk.Frame(api_frame)
            row.pack(fill='x', padx=10, pady=4)
            ttk.Label(row, text=label, width=12).pack(side='left')
            ttk.Entry(row, textvariable=var, width=width, show='*' if secret else '').pack(
                side='left', fill='x', expand=True
            )

        api_row('API 주소', self.api_base_var)
        api_row('업체 코드', self.tenant_slug_var, width=20)
        api_row('아이디', self.api_email_var, width=28)
        api_row('비밀번호', self.api_password_var, secret=True, width=28)

        login_row = ttk.Frame(api_frame)
        login_row.pack(fill='x', padx=10, pady=8)
        ttk.Checkbutton(login_row, text='API 계정 저장', variable=self.remember_api_var).pack(side='left')
        ttk.Button(login_row, text='로그인', command=self._api_login).pack(side='left', padx=(12, 0))
        ttk.Button(login_row, text='매칭 서식 새로고침', command=self._api_refresh_profiles).pack(
            side='left', padx=(8, 0)
        )

        profile_row = ttk.Frame(api_frame)
        profile_row.pack(fill='x', padx=10, pady=(0, 8))
        ttk.Label(profile_row, text='매칭 서식', width=12).pack(side='left')
        self.profile_combo = ttk.Combobox(profile_row, textvariable=self.profile_var, width=40, state='readonly')
        self.profile_combo.pack(side='left', fill='x', expand=True, padx=(0, 6))
        ttk.Button(profile_row, text='타임트리 자동 생성', command=self._api_create_profile).pack(side='left')

        file_frame = ttk.LabelFrame(self, text='업로드 파일')
        file_frame.pack(fill='x', padx=12, pady=6)

        def file_row(label: str, var: tk.StringVar, browse) -> None:
            row = ttk.Frame(file_frame)
            row.pack(fill='x', padx=10, pady=4)
            ttk.Label(row, text=label, width=14).pack(side='left')
            ttk.Entry(row, textvariable=var).pack(side='left', fill='x', expand=True, padx=(0, 6))
            ttk.Button(row, text='찾기', command=browse).pack(side='left')

        file_row('매칭 결과 엑셀', self.ctx.matched_xlsx, self._browse_matched)
        file_row('업로드용 엑셀', self.ctx.upload_xlsx, self._browse_upload)

        opt_row = ttk.Frame(file_frame)
        opt_row.pack(fill='x', padx=10, pady=4)
        ttk.Label(opt_row, text='운영사', width=14).pack(side='left')
        ttk.Entry(opt_row, textvariable=self.operating_company_var, width=24).pack(side='left')
        ttk.Label(
            opt_row,
            text='(청소비서 운영사명. 업체코드 넣지 마세요. 비우면 생략)',
            foreground='#64748b',
        ).pack(side='left', padx=(6, 0))
        opt_row2 = ttk.Frame(file_frame)
        opt_row2.pack(fill='x', padx=10, pady=(0, 4))
        ttk.Checkbutton(opt_row2, text='등록가능(Y) 행만', variable=self.import_ready_only_var).pack(
            side='left', padx=(16, 0)
        )
        ttk.Checkbutton(opt_row2, text='확정(Y) 행만', variable=self.confirmed_only_var).pack(side='left', padx=(8, 0))
        ttk.Label(opt_row2, text='(등록가능·제외 필터 적용)', foreground='#64748b').pack(side='left', padx=(6, 0))

        btn_row = ttk.Frame(file_frame)
        btn_row.pack(fill='x', padx=10, pady=8)
        ttk.Button(btn_row, text='① 업로드용 엑셀 만들기', command=self._export_upload_xlsx).pack(side='left')
        ttk.Button(btn_row, text='② 미리보기', command=self._api_preview).pack(side='left', padx=(8, 0))
        ttk.Button(btn_row, text='③ 일괄 등록 실행', command=self._api_execute).pack(side='left', padx=(8, 0))
        self.resume_btn = ttk.Button(btn_row, text='③-b 이어서 등록', command=self._api_resume_execute)
        self.resume_btn.pack(side='left', padx=(8, 0))

        undo_frame = ttk.LabelFrame(self, text='일괄 등록 되돌리기')
        undo_frame.pack(fill='x', padx=12, pady=6)

        undo_row = ttk.Frame(undo_frame)
        undo_row.pack(fill='x', padx=10, pady=6)
        ttk.Label(undo_row, text='실행 이력', width=14).pack(side='left')
        self.import_run_combo = ttk.Combobox(
            undo_row, textvariable=self.import_run_var, width=52, state='readonly'
        )
        self.import_run_combo.pack(side='left', fill='x', expand=True, padx=(0, 6))
        ttk.Button(undo_row, text='이력 새로고침', command=self._api_refresh_runs).pack(side='left')

        undo_btn_row = ttk.Frame(undo_frame)
        undo_btn_row.pack(fill='x', padx=10, pady=(0, 8))
        ttk.Label(
            undo_btn_row,
            text='선택한 실행으로 등록된 접수를 비밀번호 확인 후 영구 삭제합니다.',
            foreground='#64748b',
            wraplength=720,
        ).pack(side='left', fill='x', expand=True)
        ttk.Button(undo_btn_row, text='④ 일괄 등록 삭제', command=self._api_delete_run).pack(side='right')

        self.summary_text = scrolledtext.ScrolledText(self, height=12, state='disabled', wrap='word')
        self.summary_text.pack(fill='x', padx=12, pady=8)

    def _persist_api_settings(self) -> None:
        s = self.ctx.settings
        s.api_base_url = self.api_base_var.get().strip()
        s.tenant_slug = self.tenant_slug_var.get().strip()
        s.api_email = self.api_email_var.get().strip()
        s.api_password = self.api_password_var.get()
        s.remember_api = bool(self.remember_api_var.get())
        s.operating_company = self.operating_company_var.get().strip()
        s.upload_import_ready_only = bool(self.import_ready_only_var.get())
        s.upload_confirmed_only = bool(self.confirmed_only_var.get())
        self.ctx.persist_settings()
        if self.profile_combo['values']:
            selected = self.profile_var.get().strip()
            for profile in self._profiles:
                if profile.get('name') == selected:
                    s.profile_id = str(profile.get('id', ''))
                    break
        save_settings(s)

    def _initial_dir(self) -> str:
        brand = self.ctx.settings.brand_name.strip()
        root = self.ctx.settings.workspace_root
        if brand:
            return str(brand_dir(brand, root))
        return str(resolve_workspace_root(root))

    def _browse_matched(self) -> None:
        path = filedialog.askopenfilename(
            title='매칭 결과 엑셀', filetypes=[('Excel', '*.xlsx')], initialdir=self._initial_dir()
        )
        if path:
            self.ctx.matched_xlsx.set(path)

    def _browse_upload(self) -> None:
        path = filedialog.asksaveasfilename(
            title='업로드용 엑셀', defaultextension='.xlsx', filetypes=[('Excel', '*.xlsx')], initialdir=self._initial_dir()
        )
        if path:
            self.ctx.upload_xlsx.set(path)

    def _api_client_or_warn(self) -> CbiseoClient | None:
        if self._api_client and self._api_client.token:
            return self._api_client
        messagebox.showwarning(PANEL_TITLE, '먼저 로그인하세요.')
        return None

    def _api_login(self) -> None:
        if self._worker and self._worker.is_alive():
            return
        self._persist_api_settings()
        self.ctx.set_status('로그인 중…')
        self._worker = threading.Thread(target=self._api_login_worker, daemon=True)
        self._worker.start()

    def _api_login_worker(self) -> None:
        try:
            client = CbiseoClient(self.ctx.settings.api_base_url)
            client.login(
                self.ctx.settings.tenant_slug,
                self.ctx.settings.api_email,
                self.ctx.settings.api_password,
            )
            self._api_client = client
            self._queue.put(('profiles', client.list_profiles()))
            runs_data = client.list_import_runs(limit=30)
            runs = runs_data.get('items', []) if isinstance(runs_data, dict) else []
            self._queue.put(('import_runs', {'runs': runs, 'select': self.ctx.settings.last_import_run_id}))
            self._queue.put(('summary', '로그인 성공'))
            self._queue.put(('status', '④ 로그인 완료'))
            self._queue.put(('done', (True, '로그인했습니다.')))
        except Exception as exc:
            self._queue.put(('status', f'로그인 오류: {exc}'))
            self._queue.put(('done', (False, str(exc))))

    def _api_refresh_profiles(self) -> None:
        if not self._api_client_or_warn() or (self._worker and self._worker.is_alive()):
            return
        self._worker = threading.Thread(target=self._api_profiles_worker, daemon=True)
        self._worker.start()

    def _api_profiles_worker(self) -> None:
        try:
            profiles = self._api_client.list_profiles() if self._api_client else []
            self._queue.put(('profiles', profiles))
            self._queue.put(('status', '매칭 서식 목록 갱신'))
        except Exception as exc:
            self._queue.put(('status', f'오류: {exc}'))

    def _set_profiles(self, profiles: list[dict]) -> None:
        self._profiles = profiles
        names = [str(p.get('name', '')) for p in profiles]
        self.profile_combo['values'] = names
        selected = ''
        pid = self.ctx.settings.profile_id
        if pid:
            for profile in profiles:
                if str(profile.get('id')) == pid:
                    selected = str(profile.get('name', ''))
                    break
        if not selected:
            for profile in profiles:
                if str(profile.get('name', '')).strip() == '타임트리 자동':
                    selected = '타임트리 자동'
                    break
        if not selected and names:
            selected = names[0]
        if selected:
            self.profile_var.set(selected)

    def _selected_profile_id(self) -> str | None:
        name = self.profile_var.get().strip()
        for profile in self._profiles:
            if str(profile.get('name', '')).strip() == name:
                return str(profile.get('id', ''))
        return self.ctx.settings.profile_id or None

    def _estimate_work_units(self, upload_path: Path, *, prefer_created: bool) -> int:
        if prefer_created and self._last_preview:
            created = int(self._last_preview.get('createdCount', 0))
            if created > 0:
                return created
        try:
            return max(1, count_xlsx_data_rows(upload_path))
        except Exception:
            return 1

    def _open_progress(self, *, title: str, total_units: int, status: str, mode: str) -> None:
        self._close_progress()
        self._progress_mode = mode
        self._progress_dialog = TaskProgressDialog(
            self.root,
            title=title,
            total_units=total_units,
            status=status,
        )
        self._schedule_progress_tick()

    def _schedule_progress_tick(self) -> None:
        if self._progress_tick_job:
            try:
                self.root.after_cancel(self._progress_tick_job)
            except ValueError:
                pass
            self._progress_tick_job = None
        self._progress_tick_job = self.root.after(250, self._progress_tick)

    def _progress_tick(self) -> None:
        dialog = self._progress_dialog
        if dialog is None or dialog.closed:
            self._progress_tick_job = None
            return
        if self._progress_mode == 'preview':
            dialog.tick_time_estimate(
                seconds_per_unit=0.12,
                status='미리보기 중…',
                detail='행별 등록 가능 여부를 확인하고 있습니다.',
            )
        elif self._progress_mode == 'execute':
            dialog.tick_time_estimate(
                seconds_per_unit=0.35,
                status='일괄 등록 중…',
                detail='청소비서에 접수를 등록하고 있습니다.',
            )
        elif self._progress_mode == 'delete':
            dialog.tick_time_estimate(
                seconds_per_unit=0.08,
                status='일괄 삭제 중…',
                detail='등록된 접수를 삭제하고 있습니다.',
            )
        else:
            dialog.tick_time_estimate(status='처리 중…')
        self._schedule_progress_tick()

    def _finish_progress(self, *, success: bool, message: str) -> None:
        if self._progress_tick_job:
            try:
                self.root.after_cancel(self._progress_tick_job)
            except ValueError:
                pass
            self._progress_tick_job = None
        dialog = self._progress_dialog
        if dialog and not dialog.closed:
            dialog.finish(success=success, message=message)
        self._progress_dialog = None
        self._progress_mode = ''

    def _close_progress(self) -> None:
        if self._progress_tick_job:
            try:
                self.root.after_cancel(self._progress_tick_job)
            except ValueError:
                pass
            self._progress_tick_job = None
        if self._progress_dialog and not self._progress_dialog.closed:
            self._progress_dialog.close()
        self._progress_dialog = None
        self._progress_mode = ''

    def _format_run_label(self, run: dict) -> str:
        created = int(run.get('createdCount', 0) or 0)
        remaining = int(run.get('remainingCreatedCount', created) or 0)
        deleted = int(run.get('deletedCount', 0) or 0)
        file_name = str(run.get('fileName') or '파일명 없음')
        created_at = str(run.get('createdAt') or '')[:16].replace('T', ' ')
        run_id = str(run.get('id', ''))[:8]
        status = str(run.get('status') or '')
        last_row = run.get('lastProcessedRowIndex')
        status_short = {
            'RUNNING': '중단',
            'FAILED': '실패',
            'COMPLETED': '완료',
        }.get(status, status or '—')
        if deleted > 0 and remaining == 0:
            state = '삭제완료'
        elif status in ('RUNNING', 'FAILED'):
            pending = int(run.get('pendingRowCount', 0) or 0)
            row_hint = f'{last_row}행' if last_row else ''
            state = f'{status_short}·등록{created}·남음{pending}{("·" + row_hint) if row_hint else ""}'
        elif remaining < created:
            state = f'남음 {remaining}건'
        else:
            state = f'등록 {created}건·{status_short}'
        return f'{created_at} · {file_name} · {state} · {run_id}'

    def _set_import_runs(self, runs: list[dict], *, select_run_id: str = '') -> None:
        self._import_runs = runs
        labels: list[str] = []
        label_to_id: dict[str, str] = {}
        for run in runs:
            label = self._format_run_label(run)
            labels.append(label)
            label_to_id[label] = str(run.get('id', ''))
        self.import_run_combo['values'] = labels
        target_id = select_run_id or self.ctx.settings.last_import_run_id
        selected = ''
        if target_id:
            for label, run_id in label_to_id.items():
                if run_id == target_id:
                    selected = label
                    break
        if not selected and labels:
            selected = labels[0]
        if selected:
            self.import_run_var.set(selected)
        self._update_resume_button()

    def _update_resume_button(self) -> None:
        if not hasattr(self, 'resume_btn'):
            return
        run_id, offset, can_resume = self._selected_run_resume_state()
        if can_resume and run_id:
            self.resume_btn.state(['!disabled'])
        else:
            self.resume_btn.state(['disabled'])

    def _selected_run_resume_state(self) -> tuple[str | None, int, bool]:
        run_id = self._selected_import_run_id()
        if not run_id:
            return None, 0, False
        for run in self._import_runs:
            if str(run.get('id', '')) == run_id:
                status = str(run.get('status', ''))
                can_resume = bool(run.get('canResume')) or status in ('RUNNING', 'FAILED')
                next_offset = int(run.get('nextOffset', 0) or 0)
                if run_id == self.ctx.settings.last_import_run_id:
                    saved = int(self.ctx.settings.last_import_run_offset or 0)
                    next_offset = max(next_offset, saved)
                if status == 'COMPLETED' and not run.get('canResume'):
                    can_resume = False
                return run_id, next_offset, can_resume
        saved_id = self.ctx.settings.last_import_run_id
        if saved_id:
            return saved_id, int(self.ctx.settings.last_import_run_offset or 0), True
        return run_id, 0, False

    def _persist_resume_state(self, run_id: str, next_offset: int) -> None:
        self._persist_last_run(run_id)
        self.ctx.settings.last_import_run_offset = max(0, int(next_offset))
        save_settings(self.ctx.settings)

    def _selected_import_run_id(self) -> str | None:
        label = self.import_run_var.get().strip()
        for run in self._import_runs:
            if self._format_run_label(run) == label:
                return str(run.get('id', ''))
        return self.ctx.settings.last_import_run_id or None

    def _selected_import_run_remaining(self) -> int:
        label = self.import_run_var.get().strip()
        for run in self._import_runs:
            if self._format_run_label(run) == label:
                remaining = run.get('remainingCreatedCount')
                if remaining is None:
                    return int(run.get('createdCount', 0) or 0)
                return int(remaining or 0)
        return 0

    def _persist_last_run(self, run_id: str) -> None:
        run_id = run_id.strip()
        if not run_id:
            return
        self.ctx.settings.last_import_run_id = run_id
        save_settings(self.ctx.settings)

    def _api_refresh_runs(self) -> None:
        if not self._api_client_or_warn() or (self._worker and self._worker.is_alive()):
            return
        self._worker = threading.Thread(target=self._api_runs_worker, daemon=True)
        self._worker.start()

    def _api_runs_worker(self) -> None:
        try:
            data = self._api_client.list_import_runs(limit=30) if self._api_client else {'items': []}
            items = data.get('items') if isinstance(data, dict) else []
            runs = items if isinstance(items, list) else []
            self._queue.put(('import_runs', {'runs': runs, 'select': self.ctx.settings.last_import_run_id}))
            self._queue.put(('status', f'실행 이력 {len(runs)}건'))
        except Exception as exc:
            self._queue.put(('status', f'이력 오류: {exc}'))

    def _api_delete_run(self) -> None:
        client = self._api_client_or_warn()
        run_id = self._selected_import_run_id()
        if not client or not run_id:
            messagebox.showwarning(PANEL_TITLE, '로그인 후 삭제할 실행 이력을 선택하세요.')
            return
        remaining = self._selected_import_run_remaining()
        if remaining <= 0:
            messagebox.showinfo(PANEL_TITLE, '선택한 실행으로 남아 있는 등록 접수가 없습니다.')
            return
        if not messagebox.askyesno(
            PANEL_TITLE,
            f'선택한 일괄 등록({remaining}건)을 영구 삭제합니다.\n'
            '삭제한 접수는 복구할 수 없습니다. 계속할까요?',
        ):
            return
        password = PasswordConfirmDialog.ask(
            self.root,
            title='비밀번호 확인',
            message='본인 비밀번호를 입력하면 선택한 실행으로 등록된 접수가 삭제됩니다.',
            initial=self.ctx.settings.api_password,
        )
        if not password:
            return
        if self._worker and self._worker.is_alive():
            return
        self._persist_api_settings()
        self._open_progress(
            title='일괄 삭제',
            total_units=max(1, remaining),
            status='삭제 준비 중…',
            mode='delete',
        )
        self._worker = threading.Thread(
            target=self._api_delete_run_worker,
            args=(run_id, password),
            daemon=True,
        )
        self._worker.start()

    def _api_delete_run_worker(self, run_id: str, password: str) -> None:
        try:
            result = (
                self._api_client.delete_import_run_inquiries(run_id, password)
                if self._api_client
                else {}
            )
            deleted = int(result.get('deletedCount', 0) or 0)
            attempted = int(result.get('attemptedCount', 0) or 0)
            self._queue.put(('progress_finish', (deleted > 0, '삭제 완료' if deleted > 0 else '삭제 0건')))
            self._queue.put(('summary', format_delete_summary(result)))
            self._queue.put(('status', f"④ 삭제 — {deleted}건"))
            if self._api_client:
                runs_data = self._api_client.list_import_runs(limit=30)
                runs = runs_data.get('items', []) if isinstance(runs_data, dict) else []
                self._queue.put(('import_runs', {'runs': runs, 'select': run_id}))
            if deleted > 0:
                self._queue.put(('done', (True, f'삭제 {deleted}건 완료')))
            elif attempted <= 0:
                self._queue.put(('done', (False, '삭제할 등록 접수가 이력에 없습니다.')))
            else:
                self._queue.put(
                    (
                        'done',
                        (
                            False,
                            '삭제 0건 — 이력과 DB가 맞지 않거나 다른 실행으로 등록된 접수일 수 있습니다. '
                            '아래 요약을 확인하세요.',
                        ),
                    ),
                )
        except Exception as exc:
            self._queue.put(('progress_finish', (False, '삭제 실패')))
            self._queue.put(('done', (False, str(exc))))

    def _api_create_profile(self) -> None:
        client = self._api_client_or_warn()
        if not client:
            return
        upload_path = Path(self.ctx.upload_xlsx.get().strip())
        if not upload_path.is_file():
            messagebox.showwarning(PANEL_TITLE, '먼저 업로드용 엑셀을 만드세요.')
            return
        if self._worker and self._worker.is_alive():
            return
        self._worker = threading.Thread(target=self._api_create_profile_worker, args=(upload_path,), daemon=True)
        self._worker.start()

    def _api_create_profile_worker(self, upload_path: Path) -> None:
        try:
            from openpyxl import load_workbook
            from cbiseo_export import build_default_mapping_spec

            wb = load_workbook(upload_path, read_only=True)
            ws = wb.active
            headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            wb.close()
            spec = build_default_mapping_spec(headers)
            profile = self._api_client.find_or_create_timetree_profile(spec) if self._api_client else {}
            profiles = self._api_client.list_profiles() if self._api_client else []
            self._queue.put(('profiles', profiles))
            self._queue.put(('summary', f"매칭 서식: {profile.get('name', '')}"))
            self._queue.put(('done', (True, '매칭 서식을 준비했습니다.')))
        except Exception as exc:
            self._queue.put(('done', (False, str(exc))))

    def _export_upload_xlsx(self) -> None:
        if self._worker and self._worker.is_alive():
            return
        matched = Path(self.ctx.matched_xlsx.get().strip())
        if not matched.is_file():
            messagebox.showerror(PANEL_TITLE, f'매칭 결과 파일 없음: {matched}')
            return
        upload_raw = self.ctx.upload_xlsx.get().strip()
        upload_path = Path(upload_raw) if upload_raw else matched.with_name(matched.stem + '_청소비서업로드.xlsx')
        if not upload_raw:
            self.ctx.upload_xlsx.set(str(upload_path))
        self._persist_api_settings()
        self._worker = threading.Thread(target=self._export_upload_worker, args=(matched, upload_path), daemon=True)
        self._worker.start()

    def _export_upload_worker(self, matched: Path, upload_path: Path) -> None:
        try:
            s = self.ctx.settings
            operating, oc_warn = sanitize_operating_company(s.operating_company, s.tenant_slug)
            result = export_cbiseo_workbook(
                matched,
                upload_path,
                import_ready_only=s.upload_import_ready_only,
                confirmed_only=s.upload_confirmed_only,
                operating_company=operating,
            )
            lines = format_export_summary(result).splitlines()
            if oc_warn:
                lines.insert(1, oc_warn)
            self._queue.put(('summary', '\n'.join(lines)))
            self._queue.put(('status', f"④ 업로드용 엑셀 {result['exported_rows']}건 준비"))
            if result['exported_rows'] == 0:
                self._queue.put(('done', (False, '업로드 대상 0건입니다. 아래 안내를 확인하세요.')))
            else:
                self._queue.put(('done', (True, f"{result['exported_rows']}건 업로드용 엑셀을 만들었습니다.")))
        except Exception as exc:
            self._queue.put(('done', (False, str(exc))))

    def _api_preview(self) -> None:
        client = self._api_client_or_warn()
        profile_id = self._selected_profile_id()
        upload_path = Path(self.ctx.upload_xlsx.get().strip())
        if not client or not profile_id:
            messagebox.showwarning(PANEL_TITLE, '로그인 후 매칭 서식을 선택하세요.')
            return
        if not upload_path.is_file():
            messagebox.showerror(PANEL_TITLE, '업로드용 엑셀을 먼저 만드세요.')
            return
        self._persist_api_settings()
        total = self._estimate_work_units(upload_path, prefer_created=False)
        self._open_progress(
            title='미리보기',
            total_units=total,
            status='미리보기 준비 중…',
            mode='preview',
        )
        self._worker = threading.Thread(target=self._api_preview_worker, args=(profile_id, upload_path), daemon=True)
        self._worker.start()

    def _api_preview_worker(self, profile_id: str, upload_path: Path) -> None:
        try:
            result = self._api_client.preview_import(profile_id, upload_path) if self._api_client else {}
            self._last_preview = {
                'totalRows': int(result.get('totalRows', 0) or 0),
                'createdCount': int(result.get('createdCount', 0) or 0),
                'skippedCount': int(result.get('skippedCount', 0) or 0),
                'errorCount': int(result.get('errorCount', 0) or 0),
            }
            self._queue.put(('progress_finish', (True, '미리보기 완료')))
            self._queue.put(('summary', format_preview_summary(result)))
            self._queue.put(('done', (True, '미리보기를 완료했습니다.')))
        except Exception as exc:
            self._queue.put(('progress_finish', (False, '미리보기 실패')))
            self._queue.put(('done', (False, str(exc))))

    def _api_execute(self) -> None:
        self._start_import_execute(resume=False)

    def _api_resume_execute(self) -> None:
        run_id, offset, can_resume = self._selected_run_resume_state()
        if not can_resume or not run_id:
            messagebox.showinfo(PANEL_TITLE, '이어서 등록할 중단·실패 이력이 없습니다.')
            return
        if not messagebox.askyesno(
            PANEL_TITLE,
            f'선택한 실행({run_id[:8]}…)을 {offset + 2}행부터 이어서 등록합니다.\n'
            '이미 등록된 행·중복 접수는 자동으로 건너뜁니다.',
        ):
            return
        self._start_import_execute(resume=True)

    def _start_import_execute(self, *, resume: bool) -> None:
        client = self._api_client_or_warn()
        profile_id = self._selected_profile_id()
        upload_path = Path(self.ctx.upload_xlsx.get().strip())
        if not client or not profile_id:
            messagebox.showwarning(PANEL_TITLE, '로그인 후 매칭 서식을 선택하세요.')
            return
        if not upload_path.is_file():
            messagebox.showerror(PANEL_TITLE, '업로드용 엑셀을 먼저 만드세요.')
            return
        if not resume and not messagebox.askyesno(
            PANEL_TITLE,
            '청소비서에 접수를 일괄 등록합니다. 미리보기를 확인했나요?',
        ):
            return
        run_id, start_offset, _ = self._selected_run_resume_state() if resume else (None, 0, False)
        self._persist_api_settings()
        total = self._estimate_work_units(upload_path, prefer_created=True)
        self._open_progress(
            title='이어서 등록' if resume else '일괄 등록',
            total_units=total,
            status='등록 준비 중…',
            mode='execute',
        )
        self._worker = threading.Thread(
            target=self._api_execute_worker,
            args=(profile_id, upload_path, run_id if resume else None, start_offset if resume else 0),
            daemon=True,
        )
        self._worker.start()

    def _api_execute_worker(
        self,
        profile_id: str,
        upload_path: Path,
        resume_run_id: str | None,
        start_offset: int,
    ) -> None:
        try:
            def on_progress(batch: dict) -> None:
                self._queue.put(('progress_update', batch))

            result = (
                self._api_client.execute_import_batched(
                    profile_id,
                    upload_path,
                    run_id=resume_run_id,
                    start_offset=start_offset,
                    on_progress=on_progress,
                )
                if self._api_client
                else {}
            )
            run_id = str(result.get('runId', '')).strip()
            next_offset = int(result.get('nextOffset', 0) or 0)
            if run_id:
                self._queue.put(('resume_state', (run_id, next_offset)))
            done = bool(result.get('done'))
            status = str(result.get('status', ''))
            if done and status == 'COMPLETED':
                self._queue.put(('progress_finish', (True, '등록 완료')))
                self._queue.put(('done', (True, f"등록 {result.get('createdCount', 0)}건 완료")))
            else:
                self._queue.put(('progress_finish', (False, '일부만 등록됨')))
                self._queue.put(
                    (
                        'done',
                        (
                            False,
                            f"등록 {result.get('createdCount', 0)}건까지 진행 — "
                            f"{result.get('lastProcessedRowIndex', '?')}행에서 중단. "
                            '「이어서 등록」으로 계속하세요.',
                        ),
                    ),
                )
            self._queue.put(('summary', format_execute_summary(result)))
            self._queue.put(('status', f"④ 등록 — {result.get('createdCount', 0)}건"))
            if self._api_client and run_id:
                runs_data = self._api_client.list_import_runs(limit=30)
                runs = runs_data.get('items', []) if isinstance(runs_data, dict) else []
                self._queue.put(('import_runs', {'runs': runs, 'select': run_id}))
        except Exception as exc:
            self._queue.put(('progress_finish', (False, '등록 실패')))
            hint = (
                f'{exc}\n\n'
                '서버에 부분 이력이 저장됐을 수 있습니다. '
                '「이력 새로고침」 후 「이어서 등록」을 시도하세요.'
            )
            self._queue.put(('summary', hint))
            self._queue.put(('done', (False, str(exc))))
            if self._api_client:
                try:
                    runs_data = self._api_client.list_import_runs(limit=5)
                    runs = runs_data.get('items', []) if isinstance(runs_data, dict) else []
                    if runs:
                        self._queue.put(('import_runs', {'runs': runs, 'select': str(runs[0].get('id', ''))}))
                except Exception:
                    pass

    def _process_queue(self) -> None:
        try:
            while True:
                kind, payload = self._queue.get_nowait()
                if kind == 'status':
                    self.ctx.set_status(str(payload))
                elif kind == 'summary':
                    self.summary_text.configure(state='normal')
                    self.summary_text.delete('1.0', 'end')
                    self.summary_text.insert('1.0', str(payload))
                    self.summary_text.configure(state='disabled')
                elif kind == 'profiles':
                    self._set_profiles(payload)  # type: ignore[arg-type]
                elif kind == 'import_runs':
                    if isinstance(payload, dict) and 'runs' in payload:
                        self._set_import_runs(
                            payload['runs'],  # type: ignore[arg-type]
                            select_run_id=str(payload.get('select', '')),
                        )
                    else:
                        self._set_import_runs(payload)  # type: ignore[arg-type]
                    self._update_resume_button()
                elif kind == 'last_run':
                    self._persist_last_run(str(payload))
                elif kind == 'resume_state':
                    run_id, next_offset = payload  # type: ignore[misc]
                    self._persist_resume_state(str(run_id), int(next_offset))
                elif kind == 'progress_update':
                    batch = payload  # type: ignore[assignment]
                    if isinstance(batch, dict) and self._progress_dialog and not self._progress_dialog.closed:
                        total = int(batch.get('totalRows', 0) or 1)
                        next_offset = int(batch.get('nextOffset', 0) or 0)
                        processed = min(total, next_offset if next_offset > 0 else int(batch.get('processedRowCount', 0) or 0))
                        last_row = batch.get('lastProcessedRowIndex')
                        detail = f'등록 {batch.get("createdCount", 0)} · 건너뜀 {batch.get("skippedCount", 0)} · 오류 {batch.get("errorCount", 0)}'
                        if last_row:
                            detail += f' · 마지막 {last_row}행'
                        self._progress_dialog.set_units(
                            processed,
                            status='일괄 등록 중…',
                            detail=detail,
                        )
                        if total > 0 and processed >= total:
                            self._progress_dialog.set_units(total, status='마무리 중…', detail=detail)
                elif kind == 'progress_finish':
                    ok, message = payload  # type: ignore[misc]
                    self._finish_progress(success=bool(ok), message=str(message))
                elif kind == 'done':
                    ok, message = payload  # type: ignore[misc]
                    if ok:
                        messagebox.showinfo(PANEL_TITLE, str(message))
                    else:
                        messagebox.showerror(PANEL_TITLE, str(message))
        except queue.Empty:
            pass
        self.root.after(100, self._process_queue)
